"""
Excel 生成器
============
生成技术对比表和实验数据表，便于后续编辑。

核心类：
- ExcelGenerator：Excel 生成器
"""

import os
import logging
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """
    Excel 生成器
    
    生成两个工作表：
    1. 技术对比表：现有技术与本发明方案对比
    2. 实验数据表：供后续填写实验数据
    """
    
    def __init__(self, output_dir: str):
        """
        初始化
        
        参数：
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate(self, state: Dict[str, Any]) -> str:
        """
        生成 Excel 文件
        
        参数：
            state: 执行状态（含检索结果）
        
        返回：
            str: Excel 文件路径
        """
        wb = Workbook()
        
        # 第一个工作表：技术对比表
        ws = wb.active
        ws.title = "技术对比表"
        headers = ["项目", "现有技术/专利", "本发明方案", "优势说明"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # 从检索结果填充对比表
        retrieval_results = state.get("retrieval_results", {})
        row_idx = 2
        for key, data in retrieval_results.items():
            for item in data.get("results", []):
                ws.cell(row=row_idx, column=1, value=f"{data['tool']}-{data['query']}")
                ws.cell(row=row_idx, column=2, value=item.get("title", ""))
                ws.cell(row=row_idx, column=3, value="（待补充）")
                ws.cell(row=row_idx, column=4, value="（待补充）")
                row_idx += 1
        
        # 设置列宽
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 30
        
        # 第二个工作表：实验数据
        ws2 = wb.create_sheet("实验数据")
        ws2.append(["参数", "数值", "单位", "来源"])
        ws2.append(["示例参数", 0, "", "（请根据实际数据填写）"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        
        # 保存
        path = os.path.join(self.output_dir, "技术对比表.xlsx")
        wb.save(path)
        logger.info(f"Excel 已保存: {path}")
        return path
